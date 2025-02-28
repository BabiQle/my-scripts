import openpyxl
from collections import defaultdict


def read_excel_and_analyze(excel_file_path):
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(filename=excel_file_path)
    sheet = wb['Sheet1']

    # 用于存储每个员工的工时数据
    employee_data = defaultdict(lambda: {'projects': defaultdict(float), 'departments': defaultdict(float)})

    # 遍历数据行
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # 跳过第一行
            continue

        # 读取数据
        name = row[0]
        hours = row[6]
        project = row[12]
        department = row[14]
        if project == '数据标注组公共事务':
            continue
        if name and hours and project and department:
            # 累计每个员工在每个项目上的工时
            employee_data[name]['projects'][project] += hours
            # 累计每个员工在每个部门上的工时
            employee_data[name]['departments'][department] += hours

    # 用于存储结果
    result = {}

    # 统计每个员工的投入最多的两个项目和最多的部门
    for name, data in employee_data.items():
        # 排序项目和部门工时数据
        sorted_projects = sorted(data['projects'].items(), key=lambda x: x[1], reverse=True)
        sorted_departments = sorted(data['departments'].items(), key=lambda x: x[1], reverse=True)

        # 获取最多的两个项目
        top_projects = sorted_projects[:2] if len(sorted_projects) >= 2 else sorted_projects
        # 获取最多的一个部门
        top_department = sorted_departments[0] if sorted_departments else None

        # 保存统计结果
        result[name] = {
            'top_projects': top_projects,
            'top_department': top_department
        }

    return result


def write_results_to_excel(results, output_path):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "统计结果"

    # 写入表头
    headers = ["姓名", "项目", "投入项目2", "部门"]
    ws.append(headers)

    # 写入统计结果
    for name, stats in results.items():
        # 获取项目和部门数据
        top_projects = stats['top_projects']
        top_department = stats['top_department']

        # 准备要写入的行
        row = [name]
        for project, hours in top_projects:
            row.append(f"{project}")
        # 如果只有一个项目，补充空白以保持列对齐
        if len(top_projects) < 2:
            row.append("")
        # 添加部门数据
        if top_department:
            row.append(f"{top_department[0]}")
        else:
            row.append("")

        ws.append(row)

    # 保存工作簿
    wb.save(output_path)
    print(f"结果已保存到 {output_path}")


if __name__ == "__main__":
    input_path = r'C:\Users\zhouyi\Desktop\月底考勤核对\人员工时详情_20250219_01.01~01.31.xlsx'
    output_path = r'C:\Users\zhouyi\Desktop\月底考勤核对\统计结果.xlsx'

    # 读取并分析数据
    result = read_excel_and_analyze(input_path)

    # 将结果写入 Excel 文件
    write_results_to_excel(result, output_path)
