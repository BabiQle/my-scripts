import openpyxl
from collections import defaultdict
import os


def process_own_department(original_value):
    """处理所属部门字段的专用函数"""
    original_value = str(original_value)
    if "合作伙伴" in original_value:
        parts = original_value.split("/")
        try:
            return f"{parts[1]}/{parts[2]}"
        except IndexError:
            return original_value
    else:
        parts = original_value.split("/")
        try:
            return f"{parts[0]}/{parts[1]}"
        except IndexError:
            return original_value


def read_excel_and_analyze(excel_file_path):
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(filename=excel_file_path)
    sheet = wb['Sheet1']

    employee_data = {}
    departments = set()

    # 遍历数据行（跳过标题行）
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:
            continue  # 跳过标题行

        # 解析数据
        name = row[0]
        hours = row[6]
        own_dp = process_own_department(row[4])  # 处理后的所属部门
        member = row[1]  # 成员岗位
        project = row[12]
        raw_department = row[14]  # 原始的department值

        # 新增处理逻辑
        if project == "公共休假项目":
            actual_department = "无所属部门"
        else:
            if raw_department is None or str(raw_department).strip() == "":
                raise ValueError(f"非公共休假项目且部门为空，行数据：{row}")
            actual_department = str(raw_department).strip()

        # 处理空值并添加到departments集合
        departments.add(actual_department)

        # 转换工时为数值类型
        try:
            hours = float(hours)
        except (ValueError, TypeError):
            hours = 0.0

        # 构建员工数据结构
        if name not in employee_data:
            employee_data[name] = {
                '所属部门': own_dp,
                '成员岗位': member,
                '工时统计': defaultdict(float)}

        # 累加工时
        employee_data[name]['工时统计'][actual_department] += hours

    # 转换数据结构以便输出
    departments = sorted(departments)
    result = {
        '所有部门': departments,
        '员工数据': {}
    }

    for name, data in employee_data.items():
        result['员工数据'][name] = {
            '所属部门': data['所属部门'],
            '成员岗位': data['成员岗位'],
            '部门工时': dict(data['工时统计'])}

    return result['所有部门'], result['员工数据']


def write_results_to_excel(all_departments, employee_data, output_path):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "统计结果"

    # 处理部门列表（排除"无所属部门"）
    filtered_departments = sorted([dept for dept in all_departments if dept != "无所属部门"])

    # 构建表头（按要求的顺序）
    fixed_headers = ["成员姓名", "所属部门", "成员岗位", "总人天"]
    department_headers = filtered_departments  # 过滤后的部门列表
    trailing_headers = ['休假', '备注']
    headers = fixed_headers + department_headers + trailing_headers

    # 写入表头
    ws.append(headers)

    # 写入数据行
    for name, data in employee_data.items():
        # 原始总工时计算
        total_hours = sum(data['部门工时'].values())
        total_days = round(total_hours / 8, 2)  # 精确总人天

        # 计算各部门显示值（四舍五入后）
        department_days = [
            round(data['部门工时'].get(dept, 0) / 8, 2)
            for dept in filtered_departments
        ]
        vacation_days = round(data['部门工时'].get("无所属部门", 0) / 8, 2)

        # 计算显示值总和与精确总人天的差异
        sum_display = sum(department_days) + vacation_days
        delta = round(total_days - sum_display, 2)

        # 误差补偿逻辑
        if delta != 0:
            # 优先调整休假列
            if vacation_days + delta >= 0:  # 确保不会出现负数
                vacation_days = round(vacation_days + delta, 2)
            else:
                # 如果休假列不足调整，调整最后一个部门
                if department_days:
                    department_days[-1] = round(department_days[-1] + delta, 2)

        # 构建数据行
        row = [
            name,
            data['所属部门'],
            data['成员岗位'],
            total_days
        ]
        row.extend(department_days)
        row += [vacation_days, '']

        ws.append(row)

    # 设置数字格式（保留两位小数）
    for row in ws.iter_rows(min_row=2):
        # 设置所有数值列格式
        for idx in range(3, len(headers)):  # 从第4列（总人天）开始
            if idx == len(headers) - 1: continue  # 跳过备注列
            row[idx].number_format = '0.00'

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # 保存工作簿
    wb.save(output_path)
    print(f"结果已保存到 {output_path}")


if __name__ == "__main__":
    try:
        # 输入文件路径
        input_path = r'C:\Users\zhouyi\Desktop\月底考勤核对\人员工时详情_20250224_12.01~12.31.xlsx'
        # input_path = input("请输入输入Excel文件的完整路径（例如：C:\\path\\to\\input.xlsx）：\n").strip()
        if not input_path:
            print("未输入文件路径，程序退出。")
            exit()

        # 默认输出路径：与输入文件同级目录，命名为“统计结果.xlsx”
        output_dir = os.path.dirname(input_path)  # 获取输入文件的目录
        output_filename = "统计结果12月.xlsx"  # 固定输出文件名
        output_path = os.path.join(output_dir, output_filename)  # 拼接完整输出路径

        # 处理数据
        departments, employee_data = read_excel_and_analyze(input_path)
        if departments and employee_data:
            write_results_to_excel(departments, employee_data, output_path)

    except Exception as e:
        print(f"程序运行出错: {str(e)}")

    # 等待用户按任意键退出
    input("按任意键退出...")