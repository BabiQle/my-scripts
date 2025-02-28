import openpyxl
from collections import defaultdict


def read_excel_and_analyze(excel_file_path):
    # 读取 Excel 文件
    wb = openpyxl.load_workbook(filename=excel_file_path)
    sheet = wb['Sheet1']

    # 用于存储每个员工的工时数据
    employee_data = defaultdict(lambda: defaultdict(float))

    # 遍历数据行
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i == 0:  # 跳过第一行
            continue

        # 读取数据
        name = row[0]
        hours = row[6]
        project = row[12]
        # if project == '数据标注组公共事务':
        if project != '香港科技大学（广州）NFF项目':
            continue
        if name and hours and project:
            # 累计每个员工在每个项目上的工时
            employee_data[name][project] += hours

    # 用于存储结果
    result = {}

    # 统计每个员工的投入最多的三个项目
    for name, projects in employee_data.items():
        # 排序项目工时数据
        sorted_projects = sorted(projects.items(), key=lambda x: x[1], reverse=True)

        # 获取最多的三个项目
        top_projects = sorted_projects[:3]

        # 保存统计结果
        result[name] = top_projects

    return result


def write_results_to_excel(results, output_path):
    # 创建一个新的工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "项目工时投入top3"

    # 写入表头
    headers = ["姓名", "投入项目1", "投入项目2", "投入项目3"]
    ws.append(headers)

    # 写入统计结果
    for name, top_projects in results.items():
        # 准备要写入的行
        row = [name]
        for project, hours in top_projects:
            row.append(f"{project}: {hours} hours")
        # 如果项目少于三个，补充空白以保持列对齐
        while len(row) < 4:
            row.append("")

        ws.append(row)

    # 保存工作簿
    wb.save(output_path)
    print(f"结果已保存到 {output_path}")


if __name__ == "__main__":
    input_path = r'C:\Users\zhouyi\Desktop\月底考勤核对\人员工时详情_20250219_01.01~01.31.xlsx'
    output_path = r'C:\Users\zhouyi\Desktop\月底考勤核对\1月项目工时投入top3.xlsx'

    # 读取并分析数据
    result = read_excel_and_analyze(input_path)

    # 将结果写入 Excel 文件
    write_results_to_excel(result, output_path)
